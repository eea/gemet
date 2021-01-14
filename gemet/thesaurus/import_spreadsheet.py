from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from django.db import transaction
from django.utils import timezone

from gemet.thesaurus import PENDING, PUBLISHED, DELETED_PENDING
from gemet.thesaurus.models import (
    Namespace, Version, Concept, Property, PropertyType, Relation, Language
)
from gemet.thesaurus.utils import get_new_code


class ImportError(Exception):
    pass


def row_dicts(sheet):
    rows = sheet.rows

    mandatory_columns = {
        "Term", "Definition", "Definition source"
    }
    optional_columns = {
        "Broader concept", "Broader URI", "Group", "Theme", "Note",
        "Target language"
    }
    supported_columns = mandatory_columns.union(optional_columns)
    column_names = [
        c.value.strip() for c in next(rows) if c.value and c.value.strip()
    ]

    for column in mandatory_columns:
        if column not in column_names:
            raise ImportError(u'Column "{}" is mandatory.'.format(column))

    for column in column_names:
        if column not in supported_columns:
            for text in ['Alt Label', 'Broader concept', 'Group', 'Theme']:
                if text in column:
                    break
            else:
                raise ImportError(
                    u'Column "{}" is not supported.'.format(column)
                )

    for row in rows:
        values = [(c.value or '').strip() for c in row[1:]]
        if not ''.join(values).strip():
            # The sheet is over, there are only empty rows now.
            return
        yield dict(zip(column_names, values))


class Importer(object):

    def __init__(self, import_obj):
        self.import_obj = import_obj

    @transaction.atomic
    def import_file(self):
        """ Imports data from file and returns string with results """
        self.concept_ns = Namespace.objects.get(heading='Concepts')
        self.group_ns = Namespace.objects.get(heading='Groups')
        # Number of regular concepts before
        num_reg_concepts_bef = Concept.objects.filter(
            namespace=self.concept_ns
        ).count()

        try:
            print("Opening file...")
            wb = load_workbook(filename=self.import_obj.spreadsheet.path)
        except InvalidFileException:
            raise ImportError('The file provided is not a valid excel file.')
        except IOError:
            raise ImportError('The file provided does not exist.')

        # Get the version with no name, used for pending concepts
        self.version = Version.under_work()
        # Keep a cache with a reference to all created concepts
        self.concepts = {}

        results = ""

        # The 'EN' sheet must have the original English concepts
        if 'EN' in wb.sheetnames:
            print('Creating concepts...')
            self._create_concepts(wb['EN'])

            # Fetch the property types needed to create relations
            self.property_types = PropertyType.objects.filter(
                name__in=['broader', 'group', 'theme']
            )

            print('Creating relations...')
            self._create_relations(wb['EN'])

            num_reg_concepts_after = Concept.objects.filter(
                namespace=self.concept_ns
            ).count()

            results = ("Created {} concepts.").format(
                num_reg_concepts_after - num_reg_concepts_bef,
            )

        # All other sheets must have translations
        translation_sheetnames = [sn for sn in wb.sheetnames if sn != 'EN']

        if translation_sheetnames:
            print('Creating translations...')
            for sheetname in translation_sheetnames:
                print('    {}...'.format(sheetname))
                self._add_translations(wb[sheetname])

            if results:
                results += '\n\n'
            results += (
                "Created translations for the following {} languages: {}."
            ).format(
                len(translation_sheetnames),
                ', '.join(translation_sheetnames)
            )

        return results

    def _create_concepts(self, sheet):
        for i, row in enumerate(row_dicts(sheet)):
            label = row.get("Term")  # aka prefLabel
            if not label:
                raise ImportError(u'Row {} has no "Term".'.format(i))

            alt_labels = [
                row[key] for key in row.keys()
                if 'Alt Label' in key and row[key]
            ]
            defin = row.get("Definition")
            source = row.get("Definition source")
            note = row.get("Note")

            property_values = {
                'prefLabel': label,
                'definition': defin,
                'source': source,
                'scopeNote': note,
            }

            if alt_labels:
                property_values['altLabel'] = alt_labels

            # A concept must always have at least an English property, so if
            # there is no English property corresponding to that term in the
            # DB, the concept must be new.
            prop = Property.objects.filter(
                name='prefLabel',
                value__iexact=label,
                language_id='en',
            ).first()

            if prop:
                concept = prop.concept
                msg = u'Concept {} exists. '.format(label)
                if prop.status in [PENDING, PUBLISHED]:
                    del property_values['prefLabel']
                    msg += 'Skipping prefLabel creation.'
                print(msg)
            else:
                code = get_new_code(self.concept_ns)

                concept = Concept.objects.create(
                    code=code,
                    namespace=self.concept_ns,
                    version_added=self.version,
                    status=PENDING,
                    date_entered=timezone.now(),
                )
                print(u'Concept added: {}'.format(label))

            self.concepts[label.lower()] = concept

            concept.update_or_create_properties(property_values)

    def _create_relations(self, sheet):

        for i, row in enumerate(row_dicts(sheet)):
            print("Processing row {}...".format(i))

            self._create_row_relations(i, row)

    def _create_row_relations(self, i, row):
        source_label = row.get("Term")  # aka prefLabel
        source = self.concepts[source_label.lower()]

        for property_type in self.property_types:

            # Look for columns specifying relationships
            if property_type.name == 'broader':
                target_labels = [
                    row[key] for key in row.keys()
                    if 'Broader concept' in key and row[key]
                ]
                broader_labels = target_labels
            elif property_type.name == 'group':
                target_labels = [
                    row[key] for key in row.keys()
                    if 'Group' in key and row[key]
                ]
            elif property_type.name == 'theme':
                target_labels = [
                    row[key] for key in row.keys()
                    if 'Theme' in key and row[key]
                ]

            if not target_labels:
                # If it doesn't exist, there is no relation to be created
                print(
                    (
                        'Row {} has no relationship columns '
                        '(i.e. "Broader concept", "Group", or "Theme").'
                    ).format(i)
                )
                return

            for target_label in target_labels:
                target = Concept.objects.filter(
                    properties__name='prefLabel',
                    properties__value=target_label,
                    properties__language_id='en',
                    namespace=property_type.namespace
                ).exclude(
                    status=DELETED_PENDING,
                    properties__status=DELETED_PENDING
                ).first()

                if not target:
                    raise ImportError(
                        'Row {}: concept "{}" does not exist.'.format(
                            i, target_label
                        )
                    )

                relation, created = Relation.objects.get_or_create(
                    source=source,  # child
                    target=target,  # parent
                    property_type=property_type,
                    defaults={
                        'version_added': self.version, 'status': PENDING
                    }
                )

                if created:
                    print('Relation created: {}'.format(relation))

                if not relation.reverse:
                    reverse_relation = relation.create_reverse()
                    print(
                        'Reverse relation created: {}'.format(
                            reverse_relation
                        )
                    )

        created = source.inherit_groups_and_themes_from_broader()
        print(
            'Inherited groups and themes from: {}'.format(
                ', '.join(broader_labels)
            )
        )

    def _add_translations(self, sheet):
        for row in row_dicts(sheet):

            language = Language.objects.get(code=sheet.title.lower())

            en_label = row.get('Term')
            if not en_label:
                raise ImportError(u'"Term" column cannot be blank.')
            translation = row.get('Target language')
            definition = row.get('Definition')
            source = row.get("Definition source")
            note = row.get("Note")
            alt_labels = [
                row[key] for key in row.keys()
                if 'Alt Label' in key and row[key]
            ]

            property_values = {
                'prefLabel': translation,
                'definition': definition,
                'source': source,
                'scopeNote': note,
            }
            if alt_labels:
                property_values['altLabel'] = alt_labels

            concept = Property.objects.get(
                name='prefLabel',
                value__iexact=en_label,
                language=Language.objects.get(code='en'),
            ).concept

            concept.update_or_create_properties(
                property_values, language_id=language.code
            )

    def _get_concept(self, label):
        concept = self.concepts.get(label.lower())

        if not concept:
            try:
                concept = Property.objects.get(
                    name='prefLabel',
                    value__iexact=label,
                    language_id='en',
                    concept__namespace=self.concept_ns,
                ).concept
            except Property.DoesNotExist:
                concept = None

        return concept
