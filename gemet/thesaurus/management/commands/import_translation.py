from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from django.core.management import CommandError
from django.core.management.base import BaseCommand

from gemet.thesaurus import DELETED_PENDING, PENDING, PUBLISHED
from gemet.thesaurus import models
from gemet.thesaurus.utils import get_new_code, get_search_text, refresh_search_text
from gemet.thesaurus.utils import split_text_into_terms


class Command(BaseCommand):
    help = (
        "Import translations for existing concepts from Excel "
        "(column A=English name, B=translated name, C=translated definition)"
    )

    def add_arguments(self, parser):
        parser.add_argument('excel_file')
        parser.add_argument('language_code')

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=options['excel_file'])
        except InvalidFileException:
            raise CommandError('The file provided is not a valid excel file.')
        except IOError:
            raise CommandError('The file provided does not exist.')

        sheet = wb.active
        langcode = options['language_code']
        try:
            self.language = models.Language.objects.get(code=langcode)
        except models.Language.DoesNotExist:
            raise CommandError('Language "{}" not found.'.format(langcode))

        self.version = models.Version.under_work()
        self.namespace = models.Namespace.objects.get(heading='Concepts')
        self.concepts = {}

        self.stdout.write('Adding translations...')
        self._add_translations(sheet)

    def _add_translations(self, sheet):
        for row in sheet.iter_rows(max_col=3, min_row=2):
            en_label, label, definition = [(cell.value or '').strip() for cell in row]

            if not en_label:
                # Skip empty rows
                continue
            properties = {
                'prefLabel': label,
                'definition': definition,
            }

            property = models.Property.objects.filter(
                name='prefLabel',
                value__iexact=en_label,
                language=models.Language.objects.get(code='en'),
            ).first()
            if not property:
                # Concept not found
                self.stderr.write(u'Skipping concept not found: {}'.format(en_label))
                continue
            concept = property.concept
            msg = u'Concept {} exists. '.format(en_label)

            for name, value in properties.iteritems():
                current_property = models.Property.objects.filter(
                    concept=concept,
                    language=self.language,
                    name=name,
                    status__in=[PENDING, PUBLISHED]
                ).first()
                if current_property:
                    if current_property.status == PENDING:
                        current_property.value = value
                        current_property.save()
                    else:
                        current_property.status = DELETED_PENDING
                        current_property.save()
                if not (current_property and
                        current_property.status == PENDING):
                    models.Property.objects.create(
                        status=PENDING,
                        version_added=self.version,
                        concept=concept,
                        language=self.language,
                        name=name,
                        value=value,
                    )
            refresh_search_text('prefLabel', concept.id, self.language.code)

