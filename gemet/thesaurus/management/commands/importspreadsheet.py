from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from django.core.management import CommandError
from django.core.management.base import BaseCommand

from gemet.thesaurus import PENDING, PUBLISHED
from gemet.thesaurus import models
from gemet.thesaurus.management.utils import get_search_text
from gemet.thesaurus.utils import get_new_code, split_text_into_terms
from gemet.thesaurus.utils import has_reverse_relation, create_reverse_relation

NAMESPACE = 'Concepts'
LANGCODE = 'en'


class Command(BaseCommand):
    help = 'Import new concepts from Excel spreadsheet'

    def add_arguments(self, parser):
        parser.add_argument('excel_file')

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=options['excel_file'])
        except InvalidFileException:
            raise CommandError('The file provided is not a valid excel file.')
        except IOError:
            raise CommandError('The file provided does not exist.')

        sheet = wb.active
        self.language = models.Language.objects.get(code=LANGCODE)
        self.version = models.Version.under_work()
        self.namespace = models.Namespace.objects.get(heading=NAMESPACE)
        self.concepts = {}
        self.properties = []

        self.stdout.write('Creating concepts...')
        self._create_concepts(sheet)
        self.stdout.write('Creating relations...')
        self._create_relations(sheet)

        self.stdout.write('Creating {} properties...'
                          .format(len(self.properties)))
        models.Property.objects.bulk_create(self.properties, batch_size=10000)

    def _create_concepts(self, sheet):
        for row in sheet.iter_rows(max_col=3, min_row=2):
            label, defin, source = [(cell.value or '').strip() for cell in row]

            if not label:
                continue

            properties = {
                'prefLabel': label,
                'definition': defin,
                'source': source,
            }
            is_new_concept = False

            property = models.Property.objects.filter(
                name='prefLabel',
                value__iexact=label,
                language=self.language,
            ).first()

            if property:
                concept = property.concept
                msg = u'Concept {} exists. '.format(label)
                if property.status in [PENDING, PUBLISHED]:
                    del properties['prefLabel']
                    msg += 'Skipping prefLabel creation.'
                self.stdout.write(msg)
            else:
                is_new_concept = True
                code = get_new_code(self.namespace)

                concept = models.Concept.objects.create(
                    code=code,
                    namespace=self.namespace,
                    version_added=self.version,
                    status=PENDING,
                )
                self.stdout.write(u'Concept added: {}'.format(label))

            self.concepts[label.lower()] = concept
            for name, value in properties.iteritems():
                if not models.Property.objects.filter(
                    concept=concept,
                    language=self.language,
                    name=name,
                    value=value,
                    status__in=[PUBLISHED, PENDING],
                ).exists():
                    is_new_concept = True
                    self.properties.append(models.Property(
                        status=PENDING,
                        version_added=self.version,
                        concept=concept,
                        language=self.language,
                        name=name,
                        value=value,
                    ))
            if is_new_concept:
                search_text = get_search_text(
                    concept.id, self.language.code, PENDING, self.version)
                if search_text:
                    self.properties.append(search_text)

    def _create_theme_group_relations(self, source):
        property_types = models.PropertyType.objects.filter(
            name__in=['theme', 'group']
        )
        property_type_broader = models.PropertyType.objects.get(name='broader')
        for property_type in property_types:
            relation = source.source_relations.filter(
                property_type=property_type).exists()
            if relation:
                self.stdout.write(
                    'Skipping {0} relation creation for concept {1}'
                        .format(property_type, source))
                continue
            broader_relations = models.Relation.objects.filter(
                property_type=property_type,
                source__target_relations__source=source,
                source__target_relations__property_type=property_type_broader
            )
            if not broader_relations:
                self.stdout.write(
                    'Skipping {0} relation creation for concept {1}.No broader.'
                        .format(property_type, source))
                continue
            for relation in broader_relations:
                new_relation = models.Relation.objects.create(
                    source=source,
                    target=relation.target,
                    property_type=property_type,
                    version_added=self.version,
                    status=PENDING)
                self.stdout.write('For concept {0} was created relation : {1}'
                                  .format(source, new_relation))

    def _create_relations(self, sheet):
        def get_terms(row, idx1, idx2):
            text = (row[idx1].value or '') + ';' + (row[idx2].value or '')
            return split_text_into_terms(text)

        related = models.PropertyType.objects.get(name='related')
        broader = models.PropertyType.objects.get(name='broader')
        narrower = models.PropertyType.objects.get(name='narrower')

        for row in sheet.iter_rows(min_row=2):
            label = (row[0].value or '').strip()

            if not label:
                continue

            relations = {
                related: get_terms(row, 3, 4),
                broader: get_terms(row, 5, 6),
                narrower: get_terms(row, 7, 8),
            }

            source = self.concepts[label.lower()]

            for property_type, terms in relations.iteritems():
                for term in terms:
                    target = self._get_concept(term)
                    if not target:
                        code = get_new_code(self.namespace)
                        target = models.Concept.objects.create(
                            code=code,
                            namespace=self.namespace,
                            version_added=self.version,
                            status=PENDING,
                        )
                        self.properties.append(models.Property(
                            status=PENDING,
                            version_added=self.version,
                            concept=target,
                            language=self.language,
                            name='prefLabel',
                            value=term,
                        ))
                        self.stdout.write('Inexistent concept: {}'.format(term))

                    relation = models.Relation.objects.filter(
                        source=source,
                        target=target,
                        property_type=property_type,
                    ).first()

                    if not relation:
                        relation = models.Relation.objects.create(
                            source=source,
                            target=target,
                            property_type=property_type,
                            version_added=self.version,
                            status=PENDING,
                        )
                        self.stdout.write('Relation created: {}'.format(relation))

                if not has_reverse_relation(relation):
                    reverse_relation = create_reverse_relation(relation)
                    self.stdout.write('Reverse relation created: {}'
                                      .format(reverse_relation))
            self._create_theme_group_relations(source)

    def _get_concept(self, label):
        concept = self.concepts.get(label.lower())

        if not concept:
            try:
                concept = (
                    models.Property.objects
                    .get(
                        name='prefLabel',
                        value__iexact=label,
                        language=self.language,
                        concept__namespace=self.namespace,
                    )
                    .concept
                )
            except models.Property.DoesNotExist:
                concept = None

        return concept
