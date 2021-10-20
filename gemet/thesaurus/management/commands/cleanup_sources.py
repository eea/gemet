from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from django.core.management import CommandError
from django.core.management.base import BaseCommand

from gemet.thesaurus import DELETED_PENDING, PENDING, PUBLISHED
from gemet.thesaurus import models


class Command(BaseCommand):
    help = (
        "Copy standalone sources to individual concepts based on status "
        "from Excel (column A=ignored, B=status, C=source code, D=count)"
    )

    def add_arguments(self, parser):
        parser.add_argument("excel_file")

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=options["excel_file"])
        except InvalidFileException:
            raise CommandError("The file provided is not a valid excel file.")
        except IOError:
            raise CommandError("The file provided does not exist.")

        sheet = wb["sources"]
        self.language = models.Language.objects.get(code="en")
        self.version = models.Version.under_work()
        self.namespace = models.Namespace.objects.get(heading="Concepts")
        self.concepts = {}

        self._cleanup_sources(sheet)

    def _cleanup_sources(self, sheet):
        for row in sheet.iter_rows(max_col=3, min_row=2):
            _descr, status, abbr = [(cell.value or "").strip() for cell in row]

            if not _descr:
                # Skip empty rows
                continue
            to_keep = status in ["OK", "no link"]
            self.stdout.write("Processing source {} (keep: {})".format(abbr, to_keep))

            source = models.DefinitionSource.objects.filter(abbr=abbr).first()
            if not source:
                # Source not found ???
                self.stderr.write(u"Skipping source not found: {}".format(abbr))
            source_descr = ", ".join(
                filter(
                    None,
                    (
                        source.url,
                        source.author,
                        source.title,
                        source.publication,
                        source.place,
                        source.year,
                    ),
                )
            )

            props = models.Property.objects.filter(
                name="source",
                status__in=[PENDING, PUBLISHED],
                value__iexact=abbr,
                language=self.language,
            ).all()
            for p in props:
                if p.status == PENDING:
                    p.value = source_descr
                    p.save()
                else:
                    # delete or replace this source with a new one
                    p.status = DELETED_PENDING
                    p.save()
                    if to_keep:
                        models.Property.objects.create(
                            status=PENDING,
                            version_added=self.version,
                            concept_id=p.concept_id,
                            language=self.language,
                            name="source",
                            value=source_descr,
                        )
            # Now delete it from definition_sources
            # source.delete()
