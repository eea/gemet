from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from django.core.management import CommandError
from django.core.management.base import BaseCommand

from gemet.thesaurus import DELETED_PENDING, PENDING, PUBLISHED
from gemet.thesaurus import models


class Command(BaseCommand):
    help = 'Update concepts English definitions from Excel spreadsheet'

    def add_arguments(self, parser):
        parser.add_argument('excel_file')

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=options['excel_file'])
        except InvalidFileException:
            raise CommandError('The file provided is not a valid excel file.')
        except IOError:
            raise CommandError('The file provided does not exist.')

        sheet = wb.active
        self.language = models.Language.objects.get(code='en')
        self.version = models.Version.under_work()
        self.namespace = models.Namespace.objects.get(heading='Concepts')

        self.stdout.write('Update concept definitions...')
        self._update_definitions(sheet)

    def _update_definitions(self, sheet):
        for row in sheet.iter_rows(max_col=4, min_row=2):
            label, new_defin, source, old_defin = [
                (cell.value or '').strip() for cell in row]

            if not label or not old_defin or old_defin == '-':
                continue

            properties = {
                'definition': new_defin,
                'source': source,
                'scopeNote': None,
            }

            property = models.Property.objects.filter(
                name='prefLabel',
                value__iexact=label,
                language=self.language,
            ).first()
            if not property:
                msg = u'Concept {} not found in db. '.format(label)
                continue

            concept = property.concept
            self.stdout.write(u'Updating concept {}.'.format(label))

            for name, value in properties.items():
                current_property = models.Property.objects.filter(
                    concept=concept,
                    language=self.language,
                    name=name,
                    status__in=[PENDING, PUBLISHED]
                ).first()

                if name == 'scopeNote':
                    old_scope = u"{}; ".format(current_property.value) if current_property else ""
                    value = (
                        u"{old_scope}Improved definition (EN only): {old_defin}"
                        .format(old_scope=old_scope, old_defin=old_defin)
                    )

                if current_property:
                    if current_property.status == PENDING:
                        current_property.value = value
                        current_property.save()
                        self.stdout.write(' Update existing {}'.format(name))
                        continue
                    else:
                        # status is PUBLISHED
                        if current_property.value == value:
                            self.stdout.write(' Skipping existing {}'.format(name))
                            continue
                        current_property.status = DELETED_PENDING
                        current_property.save()
                        msg = ' Mark as deleted the existing {}'.format(name)
                        self.stdout.write(msg)

                msg = ' Add the new {}'.format(name)
                self.stdout.write(msg)
                models.Property.objects.create(
                    status=PENDING,
                    version_added=self.version,
                    concept=concept,
                    language=self.language,
                    name=name,
                    value=value,
                )

            self.stdout.write(u' Concept updated: {}'.format(label))
