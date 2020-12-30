from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from django.core.management import CommandError
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from gemet.thesaurus import PENDING, PUBLISHED
from gemet.thesaurus import models
from gemet.thesaurus.utils import get_new_code, get_search_text


def namespace_for(property_type_name):
    heading_for = {
        'broader': 'Concepts',
        'group': 'Groups',
    }
    return models.Namespace.objects.get(
        heading=heading_for[property_type_name]
    )


def row_dicts(sheet):
    rows = sheet.rows

    mandatory_columns = {
        "Term", "Definition", "Definition reference"
    }
    optional_columns = {
        "Alt Label", "Abbreviation/Alt Label", "Synonym/Alt Label",
        "Broader concept", "Broader URI", "Group", "Note"
    }
    optional_columns.add(sheet.title)
    supported_columns = mandatory_columns.union(optional_columns)
    column_names = [
        c.value.strip() for c in next(rows) if c.value and c.value.strip()
    ]

    for column in mandatory_columns:
        if column not in column_names:
            raise CommandError(u'Column "{}" is mandatory.'.format(column))

    for column in column_names:
        if column not in supported_columns:
            raise CommandError(u'Column "{}" is not supported.'.format(column))

    for row in rows:
        values = [(c.value or '').strip() for c in row[1:]]
        if not ''.join(values).strip():
            # The sheet is over, there are only empty rows now.
            return
        yield dict(zip(column_names, values))


class Command(BaseCommand):
    """
    TODO: Convert this to a helper function and create admin action to call it
    """

    help = 'Import new concepts from Excel spreadsheet'

    def add_arguments(self, parser):
        parser.add_argument('excel_file')

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=options['excel_file'])
        except InvalidFileException:
            raise CommandError('The file provided is not a valid excel file.')
        except IOError:
            raise CommandError('The file provided does not exist.')

        # We assume all concepts being introduced are in the namespace
        # "Concepts". Adding other types of concepts is not supported.
        self.namespace_obj = models.Namespace.objects.get(heading='Concepts')
        # Get the version with no name, used for pending concepts
        self.version = models.Version.under_work()

        self.concepts = {}

        concepts_sheetname = wb.sheetnames[0]
        translation_sheetnames = wb.sheetnames[1:]

        self.stdout.write('Creating concepts...')
        self._create_concepts(wb[concepts_sheetname])
        self.stdout.write('Creating relations...')
        self._create_relations(wb[concepts_sheetname])

        if translation_sheetnames:
            self.stdout.write('Creating translations...')
            for sheetname in translation_sheetnames:
                self.stdout.write('    {}...'.format(sheetname))
                self._add_translations(wb[sheetname])

    def _create_concepts(self, sheet):
        for i, row in enumerate(row_dicts(sheet)):
            label = row.get("Term")  # aka prefLabel
            alt_labels = [row[key] for key in row.keys() if 'Alt Label' in key]
            defin = row.get("Definition")
            source = row.get("Definition source")

            if not label:
                raise CommandError(u'Row {} has no "Term".'.format(i))

            property_values = {
                'prefLabel': label,
                'definition': defin,
                'source': source,
            }

            if alt_labels:
                property_values['altLabels'] = alt_labels

            # A concept must always have at least an English property, so if
            # there is no English property corresponding to that term in the
            # DB, the concept must be new.
            prop = models.Property.objects.filter(
                name='prefLabel',
                value__iexact=label,
                language='en',
            ).first()

            if prop:
                is_new_concept = False
                concept = prop.concept
                msg = u'Concept {} exists. '.format(label)
                if prop.status in [PENDING, PUBLISHED]:
                    del property_values['prefLabel']
                    msg += 'Skipping prefLabel creation.'
                self.stdout.write(msg)
            else:
                is_new_concept = True
                code = get_new_code(self.namespace_obj)

                concept = models.Concept.objects.create(
                    code=code,
                    namespace=self.namespace_obj,
                    version_added=self.version,
                    status=PENDING,
                    date_entered=timezone.now(),
                )
                self.stdout.write(u'Concept added: {}'.format(label))

            self.concepts[label.lower()] = concept

            concept.update_or_create_properties(property_values)

            if is_new_concept:
                # Create internal "searchText" property with the concatenated
                # values from all other concept properties.
                search_text = get_search_text(
                    concept.id, 'en', PENDING, self.version
                )
                if search_text:
                    search_text.save()

    def _create_relations(self, sheet):

        property_types = models.PropertyType.objects.filter(
            name__in=['broader', 'group']
        )

        for i, row in enumerate(row_dicts(sheet)):

            source_label = row.get("Term")  # aka prefLabel

            for property_type in property_types:

                # Look for columns specifying relationships
                if property_type.name == 'broader':
                    target_label = row.get("Broader concept")
                elif property_type.name == 'group':
                    target_label = row.get("Group")

                if not target_label:
                    # If it doesn't exist, there is no relation to be created
                    self.stdout.write(
                        (
                            'Row {} has neither "broader" nor "group" columns.'
                        ).format(i)
                    )
                    continue

                source = self.concepts[source_label.lower()]
                target = self._get_concept(target_label)
                namespace = namespace_for(property_type.name)

                if not target:
                    self.stdout.write(
                        'Creating inexistent concept: {}'.format(target_label)
                    )
                    code = get_new_code(self.namespace_obj)
                    target = models.Concept.objects.create(
                        code=code,
                        namespace=namespace,
                        version_added=self.version,
                        status=PENDING,
                        date_entered=timezone.now(),
                    )
                    target.properties.create(
                        status=PENDING,
                        version_added=self.version,
                        language='en',
                        name='prefLabel',
                        value=target_label,
                    )

                # target is broader of source
                relation = models.Relation.objects.filter(
                    source=source,
                    target=target,
                    property_type=property_type,
                ).first()

                if not relation:
                    relation = models.Relation.objects.create(
                        source=source,
                        target=target,
                        property_type=property_type,
                        version_added=self.version,
                        status=PENDING,
                    )
                    self.stdout.write('Relation created: {}'.format(relation))

                if not relation.reverse:
                    reverse_relation = relation.create_reverse()
                    self.stdout.write(
                        'Reverse relation created: {}'.format(reverse_relation)
                    )

    def _add_translations(self, sheet):
        for row in row_dicts(sheet):

            language = models.Language.objects.get(code=sheet.title.lower())

            en_label = row.get('Term')
            foreign_label = row.get(sheet.title)
            definition = row.get('Definition')

            if not en_label:
                raise CommandError(u'"Term" column cannot be blank.')

            property_values = {
                'prefLabel': foreign_label,
                'definition': definition,
            }

            concept = models.Property.objects.get(
                name='prefLabel',
                value__iexact=en_label,
                language=models.Language.objects.get(code='en'),
            ).concept

            concept.update_or_create_properties(
                property_values, language_id=language.code
            )

    def _get_concept(self, label):
        concept = self.concepts.get(label.lower())

        if not concept:
            try:
                concept = models.Property.objects.get(
                    name='prefLabel',
                    value__iexact=label,
                    language='en',
                    concept__namespace=self.namespace_obj,
                ).concept
            except models.Property.DoesNotExist:
                concept = None

        return concept
