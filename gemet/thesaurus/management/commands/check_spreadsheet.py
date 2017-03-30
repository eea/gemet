from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from django.core.management import CommandError
from django.core.management.base import BaseCommand

from gemet.thesaurus.models import Property
from gemet.thesaurus.utils import split_text_into_terms


LABELS = {
    'RT (GEMET)': True,
    'RT (new)': False,
    'BT (GEMET)': True,
    'BT (new)': False,
    'NT (GEMET)': True,
    'NT (new)': False,
}


class Command(BaseCommand):
    help = 'Check if spreadsheet terms are consistent.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file')

    def check_term_existence(self, term, term_type, excel_cell_value,
                             new_terms):
        message = ' at cell {}: "{}"'.format(excel_cell_value, term)
        term_in_database = Property.objects.filter(value=term).exists()
        if term_type:
            if not term_in_database:
                message += ' not found'
                message += ' in database, but found in spreadsheet. [WARNING]' \
                    if term in new_terms else '. [ERROR]'
                self.stdout.write(message)
        else:
            if term not in new_terms:
                message += ' not found'
                message += ' in spreadsheet, but found in database. [WARNING]' \
                    if term_in_database else '. [ERROR]'
                self.stdout.write(message)

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=options['excel_file'])
        except InvalidFileException:
            raise CommandError('The file provided is not a valid excel file.')
        except IOError:
            raise CommandError('The file provided does not exist.')

        sheet = wb.active
        new_terms = [x.value.strip().lower() for x, in
                     sheet.iter_rows(min_col=1, max_col=1, min_row=2)
                     if x.value is not None]

        for label_cell, in sheet.iter_cols(min_row=1, max_row=1, min_col=1):
            if label_cell.value not in LABELS:
                continue

            term_type = LABELS.get(label_cell.value)
            for cell, in sheet.iter_rows(min_col=label_cell.col_idx,
                                         max_col=label_cell.col_idx, min_row=2):
                if not cell.value:
                    continue
                correct_terms = split_text_into_terms(cell.value)
                for term in correct_terms:
                    self.check_term_existence(term, term_type,
                                              cell.coordinate, new_terms)
